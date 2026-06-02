import hashlib
import hmac
from datetime import date
from io import BytesIO

import gspread
import pandas as pd
import plotly.express as px
import streamlit as st
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload


st.set_page_config(
    page_title="SIGE-CTAR",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =========================
# CONFIGURACIÓN
# =========================

REQUIRED_SHEETS = [
    "FACT_CTAR_SEGUIMIENTO",
    "DIM_EQUIPO",
    "DIM_SERVICIO",
    "DIM_TIPO_PROCESO",
    "DIM_ESTADO",
    "DIM_RESPONSABLE",
    "DIM_PRIORIDAD",
    "FACT_BAJAS",
    "FACT_REPOSICIONES",
    "FACT_ADQUISICIONES",
    "FACT_ALERTAS",
]

DRIVE_FOLDER_ID = "1wS6MoYjcfZGbZRtbEjPQPQTC0g_A4IGQ"
RESERVA_FILE_ID = "1YU-dxAJlkUW7BLl2DWxssHb1Fo9dmH3m"
RESERVA_WORKSHEET_NAME = "Anexo I f)"

ESTADOS_CTAR = [
    "Pendiente",
    "En gestión",
    "Requiere antecedentes",
    "En Mesa CTAR",
    "Compra iniciada",
    "OC emitida",
    "Recepcionado",
    "Finalizado",
    "Cerrado",
]

PRIORIDADES_HOSPITAL = ["🔴 Roja", "🟠 Naranjo", "🟡 Amarilla", "🟢 Verde"]

COLUMNAS_BASE_BAJAS = [
    "CARTA_SC",
    "FECHA_CARTA",
    "SIC",
    "NOMBRE_EQUIPO",
    "Serie",
    "Inventario",
    "Causal",
    "Observacion_AIF",
    "CTAR",
    "Comentario_SC",
    "Estado",
    "Presentada_a_CTAR",
]

COLUMNAS_HOSPITAL_BAJAS = [
    "PRIORIDAD_HOSPITAL",
    "JUSTIFICACION_HOSPITAL",
]

COLUMNAS_MESA_CTAR = [
    "EN_MESA_CTAR",
    "SEMANA_MESA_CTAR",
    "FECHA_MESA_CTAR",
    "OBSERVACION_MESA_CTAR",
    "ACUERDO_MESA_CTAR",
    "RESPONSABLE_ACUERDO",
]

COLUMNAS_GESTION_CTAR = [
    "GESTION_CTAR",
    "FECHA_ULTIMA_GESTION",
    "ESTADO_CTAR",
]

COLUMNAS_CIERRE = [
    "FINALIZADO",
    "CERRADO",
    "FECHA_CIERRE",
    "OBSERVACION_CIERRE",
    "FECHA_PASO_HISTORICO",
    "RESTAURAR",
    "FECHA_RESTAURACION",
    "MOTIVO_RESTAURACION",
]

COLUMNAS_CONTROL_BAJAS = (
    COLUMNAS_HOSPITAL_BAJAS
    + COLUMNAS_GESTION_CTAR
    + COLUMNAS_MESA_CTAR
    + COLUMNAS_CIERRE
)

COLUMNAS_OFICIALES_BAJAS = COLUMNAS_BASE_BAJAS + COLUMNAS_CONTROL_BAJAS

LABELS_BAJAS = {
    "CARTA_SC": "CARTA SC",
    "FECHA_CARTA": "FECHA CARTA",
    "SIC": "SIC",
    "NOMBRE_EQUIPO": "NOMBRE EQUIPO",
    "Serie": "Serie",
    "Inventario": "Inventario",
    "Causal": "Causal",
    "Observacion_AIF": "Observación AIF",
    "CTAR": "CTAR",
    "Comentario_SC": "Comentario SC",
    "Estado": "Estado",
    "Presentada_a_CTAR": "Presentada a CTAR",
    "PRIORIDAD_HOSPITAL": "PRIORIDAD HOSPITAL",
    "JUSTIFICACION_HOSPITAL": "JUSTIFICACIÓN HOSPITAL",
    "GESTION_CTAR": "GESTIÓN CTAR",
    "FECHA_ULTIMA_GESTION": "FECHA ÚLTIMA GESTIÓN",
    "ESTADO_CTAR": "ESTADO CTAR",
    "EN_MESA_CTAR": "EN MESA CTAR",
    "SEMANA_MESA_CTAR": "SEMANA MESA CTAR",
    "FECHA_MESA_CTAR": "FECHA MESA CTAR",
    "OBSERVACION_MESA_CTAR": "OBSERVACIÓN MESA CTAR",
    "ACUERDO_MESA_CTAR": "ACUERDO MESA CTAR",
    "RESPONSABLE_ACUERDO": "RESPONSABLE ACUERDO",
    "FINALIZADO": "FINALIZADO",
    "CERRADO": "CERRADO",
    "FECHA_CIERRE": "FECHA CIERRE",
    "OBSERVACION_CIERRE": "OBSERVACIÓN CIERRE",
    "FECHA_PASO_HISTORICO": "FECHA PASO HISTÓRICO",
    "RESTAURAR": "RESTAURAR",
    "FECHA_RESTAURACION": "FECHA RESTAURACIÓN",
    "MOTIVO_RESTAURACION": "MOTIVO RESTAURACIÓN",
}

# =========================
# ESTILO
# =========================

st.markdown(
    """
    <style>
    .main { background: #f4f6fb; }
    .block-container { padding-top: 1rem; }
    .header {
        background: linear-gradient(90deg, #1a2b4a, #2563eb);
        color: white;
        padding: 22px;
        border-radius: 16px;
        margin-bottom: 18px;
    }
    .header h1 { color: white; margin: 0; font-size: 38px; }
    .header p { color: #dbeafe; margin: 8px 0 0 0; font-size: 16px; }
    .metric-card {
        background: white;
        border: 1px solid #e5e7eb;
        border-radius: 16px;
        padding: 18px;
        box-shadow: 0 4px 12px rgba(15,23,42,.08);
        min-height: 112px;
    }
    .metric-title {
        font-size: 13px;
        color: #64748b;
        text-transform: uppercase;
        font-weight: 800;
        line-height: 1.25;
    }
    .metric-value {
        font-size: 34px;
        font-weight: 900;
        color: #111827;
        margin-top: 8px;
    }
    .doc-card {
        background: white;
        border: 1px solid #e5e7eb;
        border-radius: 14px;
        padding: 15px;
        margin-bottom: 12px;
        box-shadow: 0 4px 12px rgba(15,23,42,.06);
    }
    .doc-title { font-size: 16px; font-weight: 700; color: #1f2937; }
    .doc-meta { font-size: 12px; color: #64748b; margin-top: 4px; }
    .doc-link { display: inline-block; margin-top: 8px; color: #2563eb; font-weight: 700; text-decoration: none; }
    </style>
    """,
    unsafe_allow_html=True,
)

# =========================
# LOGIN
# =========================

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def get_users():
    try:
        return dict(st.secrets["auth"]["users"])
    except Exception:
        return {
            "admin": {"name": "Administrador CTAR", "password_hash": hash_password("admin123"), "role": "admin"},
            "ctar": {"name": "Usuario CTAR", "password_hash": hash_password("ctar123"), "role": "ctar"},
            "hospital": {"name": "Usuario Hospital", "password_hash": hash_password("hospital123"), "role": "hospital"},
            "ifiscal": {"name": "Inspector Fiscal", "password_hash": hash_password("if123"), "role": "if"},
        }


def authenticate(username, password):
    users = get_users()
    if username not in users:
        return None
    user = users[username]
    if hmac.compare_digest(user["password_hash"], hash_password(password)):
        return {"username": username, "name": user["name"], "role": user["role"]}
    return None


def login():
    col1, col2, col3 = st.columns([1.2, 1, 1.2])
    with col2:
        st.markdown("<h1 style='text-align:center'>🏥 SIGE-CTAR</h1>", unsafe_allow_html=True)
        st.markdown("<p style='text-align:center;color:#6b7280'>Sistema de Gestión y Trazabilidad CTAR</p>", unsafe_allow_html=True)
        with st.form("login"):
            username = st.text_input("Usuario")
            password = st.text_input("Clave", type="password")
            submit = st.form_submit_button("Ingresar", use_container_width=True)
    if submit:
        user = authenticate(username.strip(), password)
        if user:
            st.session_state["user"] = user
            st.rerun()
        else:
            st.error("Usuario o clave incorrecta.")


if "user" not in st.session_state:
    login()
    st.stop()


def role():
    return st.session_state["user"]["role"]


def can_edit():
    return role() in ["admin", "ctar", "if"]


def is_admin():
    return role() == "admin"

# =========================
# GOOGLE SHEETS / DRIVE
# =========================

def get_credentials():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    return Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scopes)


@st.cache_data(ttl=60)
def load_google_sheets():
    creds = get_credentials()
    client = gspread.authorize(creds)
    spreadsheet_id = st.secrets["google_sheet"]["spreadsheet_id"]
    spreadsheet = client.open_by_key(spreadsheet_id)
    tables = {}
    for ws in spreadsheet.worksheets():
        data = ws.get_all_records()
        df = pd.DataFrame(data)
        df.columns = limpiar_columnas_unicas(df.columns)
        tables[ws.title] = df
    return tables


def append_to_sheet(sheet_name, row_values):
    creds = get_credentials()
    client = gspread.authorize(creds)
    spreadsheet = client.open_by_key(st.secrets["google_sheet"]["spreadsheet_id"])
    worksheet = spreadsheet.worksheet(sheet_name)
    worksheet.append_row(row_values, value_input_option="USER_ENTERED")
    st.cache_data.clear()


@st.cache_data(ttl=60)
def list_drive_files(folder_id):
    creds = get_credentials()
    service = build("drive", "v3", credentials=creds)
    results = service.files().list(
        q=f"'{folder_id}' in parents and trashed = false",
        fields="files(id, name, mimeType, webViewLink, modifiedTime)",
        orderBy="modifiedTime desc",
        pageSize=100,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    return results.get("files", [])


@st.cache_data(ttl=60)
def load_reserva_presupuestaria():
    creds = get_credentials()
    service = build("drive", "v3", credentials=creds)
    request = service.files().get_media(fileId=RESERVA_FILE_ID, supportsAllDrives=True)
    file_buffer = BytesIO()
    downloader = MediaIoBaseDownload(file_buffer, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    file_buffer.seek(0)
    return pd.read_excel(file_buffer, sheet_name=RESERVA_WORKSHEET_NAME, header=None, engine="openpyxl")


def escribir_hoja(sheet_name, df_out, usar_labels=True):
    creds = get_credentials()
    client = gspread.authorize(creds)
    spreadsheet = client.open_by_key(st.secrets["google_sheet"]["spreadsheet_id"])
    try:
        worksheet = spreadsheet.worksheet(sheet_name)
    except Exception:
        worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=80)

    df_save = preparar_para_guardar(df_out, usar_labels=usar_labels)
    worksheet.clear()
    worksheet.update(
        [df_save.columns.tolist()] + df_save.astype(str).values.tolist(),
        value_input_option="USER_ENTERED",
    )
    st.cache_data.clear()

# =========================
# UTILIDADES
# =========================

def limpiar_nombre_columna(col):
    col = str(col).strip().replace("\n", " ").replace("  ", " ").replace(" ", "_")
    reemplazos = {
        "Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U",
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "Ñ": "N", "ñ": "n", ".": "", ":": "", "/": "_",
    }
    for a, b in reemplazos.items():
        col = col.replace(a, b)
    while "__" in col:
        col = col.replace("__", "_")
    return col.strip("_")


def limpiar_columnas_unicas(columns):
    seen = {}
    nuevas = []
    for c in columns:
        base = limpiar_nombre_columna(c)
        if base == "":
            base = "Columna"
        if base not in seen:
            seen[base] = 0
            nuevas.append(base)
        else:
            seen[base] += 1
            nuevas.append(f"{base}_{seen[base]}")
    return nuevas


def preparar_dataframe_streamlit(df):
    out = df.copy()
    out.columns = limpiar_columnas_unicas(out.columns)
    out = out.fillna("")
    for col in out.columns:
        out[col] = out[col].apply(lambda x: "Sí" if x is True else ("No" if x is False else x))
        out[col] = out[col].astype(str).replace({"nan": "", "None": "", "NaT": ""})
    return out


def normalizar_bool(valor, default=False):
    if pd.isna(valor):
        return default
    v = str(valor).strip().lower()
    if v in ["true", "sí", "si", "s", "1", "x", "checked", "yes"]:
        return True
    if v in ["false", "no", "n", "0", "", "none", "nan"]:
        return False
    return default


def bool_a_si_no(valor):
    return "Sí" if normalizar_bool(valor) else "No"


def clean_number(value):
    value = str(value).strip().replace(",", "").replace(" ", "")
    return pd.to_numeric(value, errors="coerce")


def header(title, subtitle):
    st.markdown(
        f"""
        <div class="header">
            <h1>{title}</h1>
            <p>{subtitle}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def metric_card(title, value):
    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-title">{title}</div>
            <div class="metric-value">{value}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def icon_by_mimetype(mime_type):
    mime_type = str(mime_type).lower()
    if "pdf" in mime_type:
        return "📕"
    if "spreadsheet" in mime_type or "excel" in mime_type:
        return "📊"
    if "document" in mime_type or "word" in mime_type:
        return "📄"
    if "presentation" in mime_type or "powerpoint" in mime_type:
        return "📽️"
    if "folder" in mime_type:
        return "📁"
    return "📎"

# =========================
# MODELO CTAR / BAJAS
# =========================

def normalizar_bajas(dataframe):
    out = dataframe.copy()
    out.columns = limpiar_columnas_unicas(out.columns)

    alias = {
        "CARTA": "CARTA_SC",
        "CARTA_SC": "CARTA_SC",
        "FECHA_CA": "FECHA_CARTA",
        "FECHA_CARTA_SC": "FECHA_CARTA",
        "FECHA_CARTA": "FECHA_CARTA",
        "NOMBRE_EQUIPO": "NOMBRE_EQUIPO",
        "OBSERVACION_AIF": "Observacion_AIF",
        "PRESENTADA_A_CTAR": "Presentada_a_CTAR",
        "PRESENTA_CTAR": "Presentada_a_CTAR",
        "PRIORIDAD": "PRIORIDAD_HOSPITAL",
        "JUSTIFICACION_PRIORIDAD": "JUSTIFICACION_HOSPITAL",
        "JUSTIFICACION": "JUSTIFICACION_HOSPITAL",
        "GESTION_CTAR": "GESTION_CTAR",
        "FECHA_ULTIMA_GESTION": "FECHA_ULTIMA_GESTION",
        "ULTIMA_GESTION": "FECHA_ULTIMA_GESTION",
        "ESTADO_CTAR": "ESTADO_CTAR",
        "EN_MESA_CTAR": "EN_MESA_CTAR",
        "MESA_CTAR": "EN_MESA_CTAR",
        "SEMANA_MESA_CTAR": "SEMANA_MESA_CTAR",
        "FECHA_MESA_CTAR": "FECHA_MESA_CTAR",
        "OBSERVACION_MESA_CTAR": "OBSERVACION_MESA_CTAR",
        "ACUERDO_MESA_CTAR": "ACUERDO_MESA_CTAR",
        "RESPONSABLE_ACUERDO": "RESPONSABLE_ACUERDO",
        "FINALIZADO": "FINALIZADO",
        "CERRADO": "CERRADO",
        "FECHA_CIERRE": "FECHA_CIERRE",
        "OBSERVACION_CIERRE": "OBSERVACION_CIERRE",
        "RESTAURAR": "RESTAURAR",
        "MOTIVO_RESTAURACION": "MOTIVO_RESTAURACION",
    }

    # Compatibilidad con estructura antigua.
    alias.update({
        "ID_Baja": "CARTA_SC",
        "ID_CTAR": "CTAR",
        "Motivo_Baja": "Causal",
        "Estado_Baja": "Estado",
        "Fecha_Baja": "FECHA_CARTA",
        "Nro_Inventario": "Inventario",
        "Equipo": "NOMBRE_EQUIPO",
    })

    rename_map = {}
    for col in out.columns:
        if col in alias:
            rename_map[col] = alias[col]
    out.rename(columns=rename_map, inplace=True)
    out.columns = limpiar_columnas_unicas(out.columns)

    for col in COLUMNAS_OFICIALES_BAJAS:
        if col not in out.columns:
            out[col] = ""

    for col in out.columns:
        out[col] = out[col].fillna("")

    out["PRIORIDAD_HOSPITAL"] = out["PRIORIDAD_HOSPITAL"].replace("", "🔴 Roja")
    out["ESTADO_CTAR"] = out["ESTADO_CTAR"].replace("", "Pendiente")

    for col in ["EN_MESA_CTAR", "FINALIZADO", "CERRADO", "RESTAURAR"]:
        out[col] = out[col].apply(normalizar_bool)

    # Si viene como Cerrado/Finalizado en estado, también marcarlo.
    estado_final = out["ESTADO_CTAR"].astype(str).str.lower().isin(["finalizado", "cerrado"])
    out.loc[estado_final, "FINALIZADO"] = True
    out.loc[estado_final, "CERRADO"] = True

    extras = [c for c in out.columns if c not in COLUMNAS_OFICIALES_BAJAS]
    out = out[COLUMNAS_OFICIALES_BAJAS + extras]
    return out


def preparar_para_guardar(df, usar_labels=True):
    out = normalizar_bajas(df) if any(c in df.columns for c in COLUMNAS_OFICIALES_BAJAS) else df.copy()
    if any(c in out.columns for c in COLUMNAS_OFICIALES_BAJAS):
        columnas = COLUMNAS_OFICIALES_BAJAS + [c for c in out.columns if c not in COLUMNAS_OFICIALES_BAJAS]
        out = out[[c for c in columnas if c in out.columns]].copy()
        for col in ["EN_MESA_CTAR", "FINALIZADO", "CERRADO", "RESTAURAR"]:
            if col in out.columns:
                out[col] = out[col].apply(bool_a_si_no)
        if usar_labels:
            out.rename(columns={c: LABELS_BAJAS.get(c, c) for c in out.columns}, inplace=True)
    out = preparar_dataframe_streamlit(out)
    return out


def cargar_bajas(tables):
    return normalizar_bajas(tables.get("FACT_BAJAS", pd.DataFrame()).copy())


def cargar_historico(tables):
    return normalizar_bajas(tables.get("HISTORICO_CTAR", pd.DataFrame()).copy())


def esta_finalizado(df):
    return (
        df["FINALIZADO"].apply(normalizar_bool)
        | df["CERRADO"].apply(normalizar_bool)
        | df["ESTADO_CTAR"].astype(str).str.lower().isin(["finalizado", "cerrado"])
    )


def separar_activos_historico(df):
    data = normalizar_bajas(df)
    finalizados = esta_finalizado(data)
    historico = data[finalizados].copy()
    activos = data[~finalizados].copy()
    if not historico.empty:
        hoy = date.today().isoformat()
        historico["FINALIZADO"] = True
        historico["CERRADO"] = True
        historico["ESTADO_CTAR"] = historico["ESTADO_CTAR"].replace("", "Finalizado")
        historico.loc[historico["ESTADO_CTAR"].astype(str).str.lower().eq("pendiente"), "ESTADO_CTAR"] = "Finalizado"
        historico["FECHA_CIERRE"] = historico["FECHA_CIERRE"].replace("", hoy)
        historico["FECHA_PASO_HISTORICO"] = historico["FECHA_PASO_HISTORICO"].replace("", hoy)
    return activos, historico


def obtener_key(df):
    for candidate in ["CARTA_SC", "SIC", "Inventario", "Serie"]:
        if candidate in df.columns and df[candidate].astype(str).str.strip().ne("").any():
            return candidate
    return "CARTA_SC"


def orden_prioridad(valor):
    v = str(valor).lower()
    if "roja" in v or "rojo" in v or "🔴" in v:
        return 1
    if "naranjo" in v or "naranja" in v or "🟠" in v:
        return 2
    if "amarilla" in v or "amarillo" in v or "🟡" in v:
        return 3
    if "verde" in v or "🟢" in v:
        return 4
    return 9


def filtrar_busqueda(df, texto):
    if not texto:
        return df
    s = texto.lower()
    mask = pd.Series(False, index=df.index)
    for col in ["CARTA_SC", "SIC", "NOMBRE_EQUIPO", "Serie", "Inventario", "Causal", "Estado", "Comentario_SC"]:
        if col in df.columns:
            mask = mask | df[col].astype(str).str.lower().str.contains(s, na=False)
    return df[mask]


def column_config_bajas(editable=False):
    return {
        "CARTA_SC": st.column_config.TextColumn("CARTA SC"),
        "FECHA_CARTA": st.column_config.TextColumn("FECHA CARTA"),
        "SIC": st.column_config.TextColumn("SIC"),
        "NOMBRE_EQUIPO": st.column_config.TextColumn("NOMBRE EQUIPO"),
        "Serie": st.column_config.TextColumn("Serie"),
        "Inventario": st.column_config.TextColumn("Inventario"),
        "Causal": st.column_config.TextColumn("Causal"),
        "Observacion_AIF": st.column_config.TextColumn("Observación AIF"),
        "CTAR": st.column_config.TextColumn("CTAR"),
        "Comentario_SC": st.column_config.TextColumn("Comentario SC"),
        "Estado": st.column_config.TextColumn("Estado"),
        "Presentada_a_CTAR": st.column_config.TextColumn("Presentada a CTAR"),
        "PRIORIDAD_HOSPITAL": st.column_config.SelectboxColumn("PRIORIDAD HOSPITAL", options=PRIORIDADES_HOSPITAL),
        "JUSTIFICACION_HOSPITAL": st.column_config.TextColumn("JUSTIFICACIÓN HOSPITAL"),
        "GESTION_CTAR": st.column_config.TextColumn("GESTIÓN CTAR"),
        "FECHA_ULTIMA_GESTION": st.column_config.TextColumn("FECHA ÚLTIMA GESTIÓN"),
        "ESTADO_CTAR": st.column_config.SelectboxColumn("ESTADO CTAR", options=ESTADOS_CTAR),
        "EN_MESA_CTAR": st.column_config.CheckboxColumn("En Mesa CTAR", default=False),
        "SEMANA_MESA_CTAR": st.column_config.TextColumn("SEMANA MESA CTAR"),
        "FECHA_MESA_CTAR": st.column_config.TextColumn("FECHA MESA CTAR"),
        "OBSERVACION_MESA_CTAR": st.column_config.TextColumn("OBSERVACIÓN MESA CTAR"),
        "ACUERDO_MESA_CTAR": st.column_config.TextColumn("ACUERDO MESA CTAR"),
        "RESPONSABLE_ACUERDO": st.column_config.TextColumn("RESPONSABLE ACUERDO"),
        "FINALIZADO": st.column_config.CheckboxColumn("Finalizado", default=False),
        "CERRADO": st.column_config.CheckboxColumn("Cerrado", default=False),
        "FECHA_CIERRE": st.column_config.TextColumn("FECHA CIERRE"),
        "OBSERVACION_CIERRE": st.column_config.TextColumn("OBSERVACIÓN CIERRE"),
        "RESTAURAR": st.column_config.CheckboxColumn("Restaurar", default=False),
        "MOTIVO_RESTAURACION": st.column_config.TextColumn("MOTIVO RESTAURACIÓN"),
    }


def vista_columnas(df, columnas):
    return preparar_dataframe_streamlit(df[[c for c in columnas if c in df.columns]].copy())


def actualizar_prioridad(base_df, respuesta_df):
    base = normalizar_bajas(base_df)
    resp = normalizar_bajas(respuesta_df)
    key = obtener_key(base)
    if key not in resp.columns:
        raise ValueError("La planilla debe mantener CARTA SC, SIC, Inventario o Serie.")
    resp_small = resp[[key] + COLUMNAS_HOSPITAL_BAJAS].drop_duplicates(subset=[key], keep="last")
    merged = base.merge(resp_small, on=key, how="left", suffixes=("", "_NUEVO"))
    for col in COLUMNAS_HOSPITAL_BAJAS:
        nuevo = f"{col}_NUEVO"
        if nuevo in merged.columns:
            merged[col] = merged[nuevo].where(merged[nuevo].astype(str).str.strip() != "", merged[col])
            merged.drop(columns=[nuevo], inplace=True)
    return normalizar_bajas(merged)


def crear_planilla_hospital(df_bajas):
    base = normalizar_bajas(df_bajas)
    columnas = COLUMNAS_BASE_BAJAS + COLUMNAS_HOSPITAL_BAJAS
    salida = preparar_para_guardar(base[[c for c in columnas if c in base.columns]], usar_labels=True)

    matriz = pd.DataFrame({
        "Prioridad": [1, 2, 3, 4],
        "Color": ["🔴 Rojo", "🟠 Naranjo", "🟡 Amarillo", "🟢 Verde"],
        "Categoría": [
            "Crítico / Muy Urgente",
            "Alta Prioridad / Urgente",
            "Prioridad Media",
            "Prioridad Baja",
        ],
        "Criterio de Priorización": [
            "La ausencia del equipamiento impacta directamente la continuidad de la atención clínica, comprometiendo la prestación del servicio o la seguridad de pacientes y funcionarios.",
            "La falta del equipamiento genera limitaciones relevantes en la capacidad operativa del servicio clínico, debiendo recurrir a alternativas parciales o redistribución de recursos.",
            "La ausencia del equipamiento no detiene la prestación clínica, pero afecta la eficiencia operativa, tiempos de respuesta o capacidad de atención del servicio.",
            "La ausencia del equipamiento no genera impacto significativo en la continuidad asistencial y existen alternativas operativas disponibles mientras se gestiona su reposición.",
        ],
    })

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        salida.to_excel(writer, index=False, sheet_name="Priorizacion_Hospital")
        matriz.to_excel(writer, index=False, sheet_name="Matriz_Prioridad")

        wb = writer.book
        ws = wb["Priorizacion_Hospital"]
        wm = wb["Matriz_Prioridad"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for sheet in [ws, wm]:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 14), 70)

        prioridad_col = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "PRIORIDAD HOSPITAL":
                prioridad_col = idx
                break
        if prioridad_col:
            col_letter = ws.cell(row=1, column=prioridad_col).column_letter
            dv = DataValidation(type="list", formula1='"🔴 Roja,🟠 Naranjo,🟡 Amarilla,🟢 Verde"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}5000")

    output.seek(0)
    return output.getvalue()


def preparar_estructura_sheets(bajas, historico):
    bajas_norm = normalizar_bajas(bajas)
    historico_norm = normalizar_bajas(historico) if not historico.empty else pd.DataFrame(columns=COLUMNAS_OFICIALES_BAJAS)
    escribir_hoja("FACT_BAJAS", bajas_norm)
    escribir_hoja("HISTORICO_CTAR", historico_norm)

# =========================
# MODELO GENERAL ANTERIOR
# =========================

def check_sheets(tables):
    return [s for s in REQUIRED_SHEETS if s not in tables]


def build_model(tables):
    fact = tables.get("FACT_CTAR_SEGUIMIENTO", pd.DataFrame()).copy()
    if fact.empty:
        return fact
    for col in ["Fecha_Ingreso", "Fecha_Compromiso"]:
        if col in fact.columns:
            fact[col] = pd.to_datetime(fact[col], errors="coerce")
    joins = [
        ("DIM_EQUIPO", "ID_Equipo"),
        ("DIM_SERVICIO", "ID_Servicio"),
        ("DIM_TIPO_PROCESO", "ID_Tipo_Proceso"),
        ("DIM_ESTADO", "ID_Estado"),
        ("DIM_RESPONSABLE", "ID_Responsable"),
        ("DIM_PRIORIDAD", "ID_Prioridad"),
    ]
    df = fact.copy()
    for sheet, key in joins:
        dim = tables.get(sheet, pd.DataFrame()).copy()
        if not dim.empty and key in df.columns and key in dim.columns:
            df = df.merge(dim, on=key, how="left")
    today = pd.Timestamp(date.today())
    if "Fecha_Ingreso" in df.columns:
        df["Dias_Desde_Ingreso"] = (today - df["Fecha_Ingreso"]).dt.days
    else:
        df["Dias_Desde_Ingreso"] = 0
    if "Fecha_Compromiso" in df.columns:
        df["Dias_Atraso"] = (today - df["Fecha_Compromiso"]).dt.days
        df["Vencido"] = df["Dias_Atraso"] > 0
    else:
        df["Dias_Atraso"] = 0
        df["Vencido"] = False
    for col in ["Equipo", "Servicio", "Tipo_Proceso", "Estado", "Responsable", "Prioridad", "Riesgo_Clinico", "Proxima_Accion"]:
        if col not in df.columns:
            df[col] = ""
    return df


def filters(df):
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        servicio = st.selectbox("Servicio", ["Todos"] + sorted(df["Servicio"].dropna().astype(str).unique().tolist()))
    with c2:
        tipo = st.selectbox("Tipo proceso", ["Todos"] + sorted(df["Tipo_Proceso"].dropna().astype(str).unique().tolist()))
    with c3:
        estado = st.selectbox("Estado", ["Todos"] + sorted(df["Estado"].dropna().astype(str).unique().tolist()))
    with c4:
        responsable = st.selectbox("Responsable", ["Todos"] + sorted(df["Responsable"].dropna().astype(str).unique().tolist()))
    out = df.copy()
    if servicio != "Todos": out = out[out["Servicio"] == servicio]
    if tipo != "Todos": out = out[out["Tipo_Proceso"] == tipo]
    if estado != "Todos": out = out[out["Estado"] == estado]
    if responsable != "Todos": out = out[out["Responsable"] == responsable]
    return out

# =========================
# CARGA
# =========================

try:
    tables = load_google_sheets()
except Exception as e:
    st.error("No se pudo conectar con Google Sheets.")
    st.code(str(e))
    st.stop()

missing = check_sheets(tables)
df = build_model(tables)
bajas_all = cargar_bajas(tables)
historico_all = cargar_historico(tables)
activos_all, historico_nuevo_tmp = separar_activos_historico(bajas_all)

# =========================
# MENÚ
# =========================

with st.sidebar:
    st.markdown("## 🏥 SIGE-CTAR")
    st.caption("Sistema CTAR")
    user = st.session_state["user"]
    st.markdown("---")
    st.caption(f"Usuario: {user['name']}")
    st.caption(f"Rol: {user['role']}")

    if role() == "hospital":
        pages = [
            "Seguimiento CTAR",
            "Mesa CTAR",
            "Histórico CTAR",
            "Repositorio Documental",
            "Reserva Presupuestaria",
        ]
    else:
        pages = [
            "Dashboard CTAR",
            "Bajas",
            "Priorización Hospital",
            "Mesa CTAR",
            "Gestión CTAR",
            "Histórico CTAR",
            "Seguimiento",
            "Reposiciones",
            "Adquisiciones",
            "Repositorio Documental",
            "Reserva Presupuestaria",
            "Alertas",
        ]
        if can_edit():
            pages.append("Registro")
        if is_admin():
            pages.append("Configuración")

    page = st.radio("Menú", pages)
    if st.button("Cerrar sesión"):
        st.session_state.clear()
        st.rerun()

# =========================
# PÁGINAS CTAR NUEVAS
# =========================

if page == "Dashboard CTAR":
    header("Dashboard CTAR", "Vista ejecutiva de bajas, priorización, Mesa CTAR, finalizados e histórico.")

    activos = activos_all.copy()
    historico_total = pd.concat([historico_all, historico_nuevo_tmp], ignore_index=True).drop_duplicates(subset=[obtener_key(historico_all if not historico_all.empty else bajas_all)], keep="last") if not historico_all.empty or not historico_nuevo_tmp.empty else pd.DataFrame(columns=COLUMNAS_OFICIALES_BAJAS)

    mesa = activos[activos["EN_MESA_CTAR"].apply(normalizar_bool)].copy()
    rojas = activos["PRIORIDAD_HOSPITAL"].astype(str).str.contains("Roja|Rojo|🔴", na=False, regex=True).sum()
    pendientes = activos["ESTADO_CTAR"].astype(str).str.lower().eq("pendiente").sum()
    en_gestion = activos["ESTADO_CTAR"].astype(str).str.lower().str.contains("gestión|gestion", na=False).sum()

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1: metric_card("📋 Casos activos", len(activos))
    with c2: metric_card("🔴 Casos críticos", int(rojas))
    with c3: metric_card("🎯 En Mesa CTAR", len(mesa))
    with c4: metric_card("✅ Finalizados", len(historico_total))
    with c5: metric_card("⏳ Pendientes", int(pendientes))

    st.markdown("---")
    g1, g2 = st.columns([1, 1])
    with g1:
        st.markdown("### Casos por prioridad")
        pr = activos.groupby("PRIORIDAD_HOSPITAL").size().reset_index(name="Cantidad")
        if not pr.empty:
            fig = px.bar(pr, x="PRIORIDAD_HOSPITAL", y="Cantidad", text="Cantidad")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay casos activos.")
    with g2:
        st.markdown("### Casos por estado CTAR")
        es = activos.groupby("ESTADO_CTAR").size().reset_index(name="Cantidad")
        if not es.empty:
            fig = px.pie(es, names="ESTADO_CTAR", values="Cantidad", hole=0.45)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay estados disponibles.")

    st.markdown("---")
    st.markdown("### 🎯 Casos programados para Mesa CTAR")
    columnas_mesa = ["SIC", "NOMBRE_EQUIPO", "Serie", "Inventario", "PRIORIDAD_HOSPITAL", "ESTADO_CTAR", "SEMANA_MESA_CTAR", "FECHA_MESA_CTAR", "ACUERDO_MESA_CTAR"]
    st.dataframe(vista_columnas(mesa, columnas_mesa), use_container_width=True, hide_index=True, column_config=column_config_bajas())

elif page == "Bajas":
    header("Bajas", "Vista general de registros activos de bajas. Desde esta sección puedes preparar la estructura de Google Sheets.")

    if is_admin():
        if st.button("🧩 Preparar/actualizar estructura FACT_BAJAS e HISTORICO_CTAR en Google Sheets"):
            preparar_estructura_sheets(bajas_all, historico_all)
            st.success("Estructura actualizada correctamente en Google Sheets.")
            st.rerun()

    activos = activos_all.copy()
    activos["ORDEN_PRIORIDAD"] = activos["PRIORIDAD_HOSPITAL"].apply(orden_prioridad)
    activos = activos.sort_values(["ORDEN_PRIORIDAD", "FECHA_CARTA"]).drop(columns=["ORDEN_PRIORIDAD"])

    c1, c2, c3, c4 = st.columns(4)
    with c1: st.metric("Activos", len(activos))
    with c2: st.metric("🔴 Rojas", int(activos["PRIORIDAD_HOSPITAL"].str.contains("Roja|Rojo", na=False, regex=True).sum()))
    with c3: st.metric("🎯 En Mesa", int(activos["EN_MESA_CTAR"].apply(normalizar_bool).sum()))
    with c4: st.metric("✅ Histórico", len(historico_all))

    search = st.text_input("Buscar por carta, SIC, equipo, serie, inventario o estado")
    vista = filtrar_busqueda(activos, search)
    columnas = COLUMNAS_BASE_BAJAS + COLUMNAS_HOSPITAL_BAJAS + COLUMNAS_GESTION_CTAR + COLUMNAS_MESA_CTAR
    st.dataframe(vista_columnas(vista, columnas), use_container_width=True, hide_index=True, column_config=column_config_bajas())

elif page == "Priorización Hospital":
    header("Priorización Hospital", "Descarga la planilla para que el Hospital priorice y luego carga la respuesta.")

    activos = activos_all.copy()
    archivo_excel = crear_planilla_hospital(activos)
    st.download_button(
        "📥 Descargar planilla de priorización para Hospital",
        data=archivo_excel,
        file_name=f"Priorizacion_Hospital_CTAR_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("### 📥 Cargar respuesta del Hospital")
    archivo = st.file_uploader("Subir planilla respondida por Hospital", type=["xlsx"], key="upload_respuesta_hospital_bajas")
    if archivo is not None:
        try:
            try:
                respuesta = pd.read_excel(archivo, sheet_name="Priorizacion_Hospital", engine="openpyxl")
            except Exception:
                archivo.seek(0)
                respuesta = pd.read_excel(archivo, sheet_name=0, engine="openpyxl")
            actualizada = actualizar_prioridad(bajas_all, respuesta)
            st.success("Planilla procesada correctamente. Revisa la vista previa antes de guardar.")
            st.dataframe(vista_columnas(actualizada, COLUMNAS_BASE_BAJAS + COLUMNAS_HOSPITAL_BAJAS), use_container_width=True, hide_index=True, column_config=column_config_bajas())
            if can_edit() and st.button("💾 Guardar priorización en FACT_BAJAS"):
                escribir_hoja("FACT_BAJAS", actualizada)
                st.success("FACT_BAJAS actualizado correctamente.")
                st.rerun()
        except Exception as e:
            st.error("No se pudo procesar la respuesta del Hospital.")
            st.code(str(e))

elif page == "Gestión CTAR":
    header("Gestión CTAR", "Admin/CTAR selecciona casos para Mesa CTAR, registra acuerdos y finaliza casos.")

    if not can_edit():
        st.error("Tu perfil no tiene permiso para editar gestión CTAR.")
        st.stop()

    activos = activos_all.copy()
    activos["ORDEN_PRIORIDAD"] = activos["PRIORIDAD_HOSPITAL"].apply(orden_prioridad)
    activos = activos.sort_values(["ORDEN_PRIORIDAD", "FECHA_CARTA"]).drop(columns=["ORDEN_PRIORIDAD"])

    search = st.text_input("Buscar por carta, SIC, equipo, serie, inventario o estado")
    vista = filtrar_busqueda(activos, search)

    columnas_editor = COLUMNAS_BASE_BAJAS + COLUMNAS_HOSPITAL_BAJAS + COLUMNAS_GESTION_CTAR + COLUMNAS_MESA_CTAR + ["FINALIZADO", "CERRADO", "FECHA_CIERRE", "OBSERVACION_CIERRE"]
    vista = vista[[c for c in columnas_editor if c in vista.columns]].copy()
    disabled_cols = [c for c in vista.columns if c not in COLUMNAS_HOSPITAL_BAJAS + COLUMNAS_GESTION_CTAR + COLUMNAS_MESA_CTAR + ["FINALIZADO", "CERRADO", "FECHA_CIERRE", "OBSERVACION_CIERRE"]]

    edited = st.data_editor(
        vista,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        disabled=disabled_cols,
        column_config=column_config_bajas(editable=True),
        key="editor_gestion_ctar_principal",
    )

    mesa_preview = normalizar_bajas(edited)
    mesa_preview = mesa_preview[mesa_preview["EN_MESA_CTAR"].apply(normalizar_bool)].copy()
    st.markdown("---")
    st.markdown("## 🎯 Vista previa: casos programados para Mesa CTAR")
    columnas_mesa = ["SIC", "NOMBRE_EQUIPO", "Serie", "Inventario", "PRIORIDAD_HOSPITAL", "ESTADO_CTAR", "SEMANA_MESA_CTAR", "FECHA_MESA_CTAR", "OBSERVACION_MESA_CTAR", "ACUERDO_MESA_CTAR", "RESPONSABLE_ACUERDO"]
    st.dataframe(vista_columnas(mesa_preview, columnas_mesa), use_container_width=True, hide_index=True, column_config=column_config_bajas())

    if st.button("💾 Guardar gestión CTAR y mover finalizados a histórico"):
        edited = normalizar_bajas(edited)
        original = normalizar_bajas(bajas_all)
        key = obtener_key(original)

        actualizado = original.copy()
        for _, row in edited.iterrows():
            row_key = str(row.get(key, ""))
            if row_key.strip() == "":
                continue
            mask = actualizado[key].astype(str) == row_key
            for col in COLUMNAS_HOSPITAL_BAJAS + COLUMNAS_GESTION_CTAR + COLUMNAS_MESA_CTAR + ["FINALIZADO", "CERRADO", "FECHA_CIERRE", "OBSERVACION_CIERRE"]:
                if col in actualizado.columns and col in edited.columns:
                    actualizado.loc[mask, col] = row[col]

        activos_final, historico_nuevo = separar_activos_historico(actualizado)
        historico_existente = normalizar_bajas(historico_all) if not historico_all.empty else pd.DataFrame(columns=COLUMNAS_OFICIALES_BAJAS)
        historico_total = pd.concat([historico_existente, historico_nuevo], ignore_index=True).fillna("")
        if key in historico_total.columns:
            historico_total = historico_total.drop_duplicates(subset=[key], keep="last")

        escribir_hoja("FACT_BAJAS", activos_final)
        escribir_hoja("HISTORICO_CTAR", historico_total)
        st.success("Gestión guardada. Los casos finalizados fueron enviados a HISTORICO_CTAR.")
        st.rerun()

elif page == "Mesa CTAR":
    header("Mesa CTAR", "Casos seleccionados para revisión de la Mesa CTAR.")

    activos = activos_all.copy()
    mesa = activos[activos["EN_MESA_CTAR"].apply(normalizar_bool)].copy()
    mesa["ORDEN_PRIORIDAD"] = mesa["PRIORIDAD_HOSPITAL"].apply(orden_prioridad)
    mesa = mesa.sort_values(["ORDEN_PRIORIDAD", "FECHA_MESA_CTAR"]).drop(columns=["ORDEN_PRIORIDAD"])

    c1, c2, c3 = st.columns(3)
    with c1: st.metric("🎯 Casos en Mesa CTAR", len(mesa))
    with c2: st.metric("🔴 Críticos en mesa", int(mesa["PRIORIDAD_HOSPITAL"].str.contains("Roja|Rojo", na=False, regex=True).sum()))
    with c3: st.metric("⏳ Pendientes en mesa", int(mesa["ESTADO_CTAR"].astype(str).str.lower().eq("pendiente").sum()))

    search = st.text_input("Buscar en Mesa CTAR")
    vista = filtrar_busqueda(mesa, search)
    columnas_mesa = ["SIC", "NOMBRE_EQUIPO", "Serie", "Inventario", "PRIORIDAD_HOSPITAL", "ESTADO_CTAR", "SEMANA_MESA_CTAR", "FECHA_MESA_CTAR", "OBSERVACION_MESA_CTAR", "ACUERDO_MESA_CTAR", "RESPONSABLE_ACUERDO", "GESTION_CTAR"]
    st.dataframe(vista_columnas(vista, columnas_mesa), use_container_width=True, hide_index=True, column_config=column_config_bajas())

elif page == "Histórico CTAR":
    header("Histórico CTAR", "Casos finalizados. Hospital visualiza trazabilidad; Administrador puede restaurar si hubo error.")

    historico_total = pd.concat([historico_all, historico_nuevo_tmp], ignore_index=True).fillna("")
    historico_total = normalizar_bajas(historico_total) if not historico_total.empty else pd.DataFrame(columns=COLUMNAS_OFICIALES_BAJAS)
    key = obtener_key(historico_total) if not historico_total.empty else "CARTA_SC"
    if key in historico_total.columns and not historico_total.empty:
        historico_total = historico_total.drop_duplicates(subset=[key], keep="last")

    c1, c2 = st.columns(2)
    with c1: st.metric("✅ Finalizados", len(historico_total))
    with c2: st.metric("🔄 Restaurables", len(historico_total) if is_admin() else 0)

    search = st.text_input("Buscar histórico por carta, SIC, equipo, serie o inventario")
    vista = filtrar_busqueda(historico_total, search)
    columnas_hist = COLUMNAS_BASE_BAJAS + COLUMNAS_HOSPITAL_BAJAS + COLUMNAS_GESTION_CTAR + COLUMNAS_MESA_CTAR + ["FECHA_CIERRE", "OBSERVACION_CIERRE"]

    if is_admin() and not vista.empty:
        st.markdown("### 🔄 Restaurar casos finalizados")
        columnas_editor = columnas_hist + ["RESTAURAR", "MOTIVO_RESTAURACION"]
        edited_hist = st.data_editor(
            vista[[c for c in columnas_editor if c in vista.columns]].copy(),
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            disabled=[c for c in columnas_editor if c not in ["RESTAURAR", "MOTIVO_RESTAURACION"]],
            column_config=column_config_bajas(),
            key="editor_historico_restaurar",
        )
        if st.button("🔄 Restaurar seleccionados a FACT_BAJAS"):
            edited_hist = normalizar_bajas(edited_hist)
            restaurar = edited_hist[edited_hist["RESTAURAR"].apply(normalizar_bool)].copy()
            if restaurar.empty:
                st.warning("No hay casos marcados para restaurar.")
            else:
                hoy = date.today().isoformat()
                restaurar["RESTAURAR"] = False
                restaurar["FINALIZADO"] = False
                restaurar["CERRADO"] = False
                restaurar["ESTADO_CTAR"] = "Pendiente"
                restaurar["FECHA_RESTAURACION"] = hoy
                activos_restaurados = pd.concat([activos_all, restaurar], ignore_index=True).fillna("")

                historico_remanente = historico_total.copy()
                key = obtener_key(historico_total)
                ids_restaurar = set(restaurar[key].astype(str))
                historico_remanente = historico_remanente[~historico_remanente[key].astype(str).isin(ids_restaurar)]

                escribir_hoja("FACT_BAJAS", activos_restaurados)
                escribir_hoja("HISTORICO_CTAR", historico_remanente)
                st.success("Casos restaurados correctamente a FACT_BAJAS.")
                st.rerun()
    else:
        st.dataframe(vista_columnas(vista, columnas_hist), use_container_width=True, hide_index=True, column_config=column_config_bajas())

elif page == "Seguimiento CTAR":
    header("Seguimiento CTAR", "Vista de solo lectura para Hospital: seguimiento de bajas, prioridades y avance de gestión CTAR.")

    activos = activos_all.copy()
    activos["ORDEN_PRIORIDAD"] = activos["PRIORIDAD_HOSPITAL"].apply(orden_prioridad)
    activos = activos.sort_values(["ORDEN_PRIORIDAD", "FECHA_CARTA"]).drop(columns=["ORDEN_PRIORIDAD"])
    mesa = activos[activos["EN_MESA_CTAR"].apply(normalizar_bool)].copy()
    historico_total = pd.concat([historico_all, historico_nuevo_tmp], ignore_index=True).fillna("")

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1: st.metric("Casos activos", len(activos))
    with c2: st.metric("🔴 Rojas", int(activos["PRIORIDAD_HOSPITAL"].str.contains("Roja|Rojo", na=False, regex=True).sum()))
    with c3: st.metric("🎯 En Mesa CTAR", len(mesa))
    with c4: st.metric("✅ Finalizados", len(historico_total))
    with c5: st.metric("⏳ Pendientes", int(activos["ESTADO_CTAR"].astype(str).str.lower().eq("pendiente").sum()))

    st.markdown("---")
    st.markdown("## 👀 Avance de gestión CTAR")
    search = st.text_input("Buscar por carta, SIC, equipo, serie, inventario o estado")
    vista = filtrar_busqueda(activos, search)
    columnas_hospital = COLUMNAS_BASE_BAJAS + COLUMNAS_HOSPITAL_BAJAS + COLUMNAS_GESTION_CTAR
    st.dataframe(vista_columnas(vista, columnas_hospital), use_container_width=True, hide_index=True, column_config=column_config_bajas())

    st.markdown("---")
    st.markdown("## 🎯 Casos programados para Mesa CTAR")
    columnas_mesa = ["SIC", "NOMBRE_EQUIPO", "Serie", "Inventario", "PRIORIDAD_HOSPITAL", "ESTADO_CTAR", "SEMANA_MESA_CTAR", "FECHA_MESA_CTAR", "OBSERVACION_MESA_CTAR", "ACUERDO_MESA_CTAR"]
    st.dataframe(vista_columnas(mesa, columnas_mesa), use_container_width=True, hide_index=True, column_config=column_config_bajas())
    st.info("Esta vista es solo de consulta. El Hospital no puede modificar información desde este perfil.")

# =========================
# PÁGINAS GENERALES EXISTENTES
# =========================

elif page == "Seguimiento":
    header("Seguimiento CTAR", "Consulta por SIC, equipo, servicio, estado, responsable y prioridad.")
    if df.empty:
        st.info("No hay datos.")
        st.stop()
    view = filters(df)
    search = st.text_input("Buscar SIC, equipo, servicio o inventario")
    if search:
        s = search.lower()
        mask = pd.Series(False, index=view.index)
        for col in ["SIC", "Equipo", "Servicio", "Nro_Inventario", "Motivo"]:
            if col in view.columns:
                mask = mask | view[col].astype(str).str.lower().str.contains(s, na=False)
        view = view[mask]
    st.dataframe(preparar_dataframe_streamlit(view), use_container_width=True, hide_index=True)

elif page == "Reposiciones":
    header("Reposiciones", "Seguimiento desde solicitud hasta compra, recepción o cierre.")
    view = df[df["Tipo_Proceso"].astype(str).str.lower().str.contains("repos", na=False)] if not df.empty else pd.DataFrame()
    st.dataframe(preparar_dataframe_streamlit(view), use_container_width=True, hide_index=True)

elif page == "Adquisiciones":
    header("Adquisiciones", "Control de procesos de compra, BACO, OC y proveedor.")
    view = df[df["Tipo_Proceso"].astype(str).str.lower().str.contains("adquis|compra", na=False)] if not df.empty else pd.DataFrame()
    adq = tables.get("FACT_ADQUISICIONES", pd.DataFrame())
    if not adq.empty and "ID_CTAR" in adq.columns and "ID_CTAR" in view.columns:
        view = view.merge(adq, on="ID_CTAR", how="left")
    st.dataframe(preparar_dataframe_streamlit(view), use_container_width=True, hide_index=True)

elif page == "Repositorio Documental":
    header("Repositorio Documental CTAR", "Consulta de documentos almacenados en Google Drive.")
    try:
        files = list_drive_files(DRIVE_FOLDER_ID)
        if not files:
            st.info("No se encontraron documentos en la carpeta de Google Drive.")
        else:
            st.success(f"Se encontraron {len(files)} documentos en el repositorio.")
            search_doc = st.text_input("Buscar documento por nombre")
            if search_doc:
                files = [f for f in files if search_doc.lower() in f.get("name", "").lower()]
            for f in files:
                icon = icon_by_mimetype(f.get("mimeType", ""))
                name = f.get("name", "Sin nombre")
                link = f.get("webViewLink", "#")
                modified = f.get("modifiedTime", "")
                st.markdown(f"""
                <div class="doc-card">
                    <div class="doc-title">{icon} {name}</div>
                    <div class="doc-meta">Última modificación: {modified}</div>
                    <a class="doc-link" href="{link}" target="_blank">Abrir documento</a>
                </div>
                """, unsafe_allow_html=True)
    except Exception as e:
        st.error("No se pudo acceder a la carpeta de Google Drive.")
        st.code(str(e))

elif page == "Reserva Presupuestaria":
    header(
        "Reserva Presupuestaria",
        "Panel ejecutivo del Anexo I f) · Fondo de reserva, desembolsos, montos en revisión y diferencia proyectada.",
    )

    def uf_fmt(valor):
        try:
            return f"{float(valor):,.2f} UF".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return "0,00 UF"

    def num_fmt(valor):
        try:
            return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return "0,00"

    def pct_fmt(valor):
        try:
            return f"{float(valor):,.1f}%".replace(".", ",")
        except Exception:
            return "0,0%"

    def comentario_financiero(diferencia, porcentaje_uso, revision, saldo):
        if diferencia < 0 and porcentaje_uso >= 80:
            return (
                "🔴 **Situación crítica:** la diferencia proyectada consolidada es negativa y el nivel de uso del fondo es elevado. "
                "Se recomienda mantener control estricto de nuevos compromisos, priorizar los casos críticos y revisar la programación financiera antes de aprobar nuevos desembolsos."
            )
        if diferencia < 0:
            return (
                "🟠 **Situación de atención:** existe una diferencia proyectada negativa. "
                "Corresponde revisar los desembolsos registrados, validar los montos en revisión y priorizar las necesidades de mayor impacto operativo."
            )
        if revision > 0:
            return (
                "🟡 **Situación en seguimiento:** el fondo presenta saldo estimado controlado, pero existen montos en revisión que podrían modificar la disponibilidad final. "
                "Se recomienda actualizar periódicamente los antecedentes y mantener seguimiento de los casos pendientes de validación."
            )
        if saldo < 0:
            return (
                "🟠 **Situación de atención:** el saldo estimado aparece negativo. "
                "Se recomienda revisar consistencia de los montos registrados y contrastar con la documentación de respaldo."
            )
        return (
            "🟢 **Situación controlada:** no se observan diferencias negativas relevantes y el uso del fondo se mantiene dentro de parámetros razonables."
        )

    def estado_anual(row):
        dif = float(row.get("Diferencia proyectada", 0) or 0)
        rev = float(row.get("En revisión desembolsos", 0) or 0)
        if dif < 0 and rev > 0:
            return "Crítico con revisión"
        if dif < 0:
            return "Déficit proyectado"
        if rev > 0:
            return "En revisión"
        return "Controlado"

    try:
        df_raw = load_reserva_presupuestaria()

        if df_raw.empty:
            st.info("No se encontraron datos en la hoja Anexo I f).")
            st.stop()

        tabla = df_raw.iloc[22:30, 1:9].copy()
        tabla.columns = [
            "Año",
            "Suplemento",
            "VMA",
            "Total general",
            "VMA año explotación",
            "Desembolsos explotación",
            "En revisión desembolsos",
            "Diferencia proyectada",
        ]

        tabla = tabla.dropna(how="all")
        tabla = tabla[tabla["Año"].astype(str).str.strip() != ""]

        columnas_uf = [
            "Suplemento",
            "VMA",
            "Total general",
            "VMA año explotación",
            "Desembolsos explotación",
            "En revisión desembolsos",
            "Diferencia proyectada",
        ]

        for col in columnas_uf:
            tabla[col] = tabla[col].apply(clean_number).fillna(0)

        fila_total = tabla[
            tabla["Año"].astype(str).str.contains("Total", case=False, na=False)
        ]

        tabla_anios = tabla[
            ~tabla["Año"].astype(str).str.contains("Total", case=False, na=False)
        ].copy()

        if fila_total.empty:
            fila_total = pd.DataFrame([{
                "VMA año explotación": tabla_anios["VMA año explotación"].sum(),
                "Desembolsos explotación": tabla_anios["Desembolsos explotación"].sum(),
                "En revisión desembolsos": tabla_anios["En revisión desembolsos"].sum(),
                "Diferencia proyectada": tabla_anios["Diferencia proyectada"].sum(),
            }])

        total_vma = fila_total["VMA año explotación"].sum()
        total_desembolsos = fila_total["Desembolsos explotación"].sum()
        total_revision = fila_total["En revisión desembolsos"].sum()
        diferencia = fila_total["Diferencia proyectada"].sum()
        saldo_disponible = total_vma - total_desembolsos - total_revision

        uso_estimado = 0
        if total_vma > 0:
            uso_estimado = ((total_desembolsos + total_revision) / total_vma) * 100

        st.markdown("## 📌 Resumen ejecutivo financiero")

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric("Fondo VMA explotación", uf_fmt(total_vma))
        with c2:
            st.metric("Desembolsado", uf_fmt(total_desembolsos))
        with c3:
            st.metric("En revisión", uf_fmt(total_revision))
        with c4:
            st.metric("Saldo estimado", uf_fmt(saldo_disponible))

        c5, c6, c7 = st.columns(3)
        with c5:
            st.metric("Diferencia proyectada", uf_fmt(diferencia))
        with c6:
            st.metric("Uso estimado del fondo", pct_fmt(uso_estimado))
        with c7:
            st.metric("Años con déficit", int((tabla_anios["Diferencia proyectada"] < 0).sum()))

        st.markdown("### Nivel de uso estimado del fondo")
        st.progress(min(max(uso_estimado / 100, 0), 1))
        st.caption(
            f"Uso estimado: **{pct_fmt(uso_estimado)}**, considerando desembolsos ejecutados y montos en revisión."
        )

        st.markdown("---")

        comentario = comentario_financiero(diferencia, uso_estimado, total_revision, saldo_disponible)
        if diferencia < 0 or saldo_disponible < 0:
            st.error(comentario)
        elif total_revision > 0:
            st.warning(comentario)
        else:
            st.success(comentario)

        st.markdown("## 🧾 Lectura ejecutiva automática")
        st.info(
            f"""
            El fondo VMA de explotación asciende a **{uf_fmt(total_vma)}**.  
            A la fecha, se registran desembolsos por **{uf_fmt(total_desembolsos)}** y montos en revisión por **{uf_fmt(total_revision)}**.  
            El saldo disponible estimado corresponde a **{uf_fmt(saldo_disponible)}**, con un uso aproximado del fondo de **{pct_fmt(uso_estimado)}**.  

            La diferencia proyectada consolidada corresponde a **{uf_fmt(diferencia)}**.
            """
        )

        st.markdown("---")
        st.markdown("## 📊 Análisis por año de explotación")

        tabla_vista = tabla.copy()
        tabla_vista["Estado financiero"] = tabla_vista.apply(estado_anual, axis=1)
        for col in columnas_uf:
            tabla_vista[col] = tabla_vista[col].apply(uf_fmt)

        st.dataframe(
            preparar_dataframe_streamlit(tabla_vista),
            use_container_width=True,
            hide_index=True,
        )

        st.markdown("---")
        st.markdown("## 📈 Comparativo financiero por año")

        grafico_base = tabla_anios.melt(
            id_vars="Año",
            value_vars=[
                "VMA año explotación",
                "Desembolsos explotación",
                "En revisión desembolsos",
            ],
            var_name="Concepto",
            value_name="UF",
        )

        fig = px.bar(
            grafico_base,
            x="Año",
            y="UF",
            color="Concepto",
            barmode="group",
            text_auto=",.0f",
            title="VMA, desembolsos y montos en revisión por año de explotación",
        )
        fig.update_layout(
            height=520,
            xaxis_title="Año de explotación",
            yaxis_title="Monto UF",
            legend_title="Concepto",
            title_x=0.02,
            bargap=0.25,
        )
        st.plotly_chart(fig, use_container_width=True)

        st.markdown("## 🔎 Diferencia proyectada por año")
        fig2 = px.bar(
            tabla_anios,
            x="Año",
            y="Diferencia proyectada",
            text="Diferencia proyectada",
            title="Diferencia proyectada respecto del fondo de reserva",
        )
        fig2.update_traces(
            texttemplate="%{text:,.0f} UF",
            textposition="outside",
        )
        fig2.update_layout(
            height=500,
            xaxis_title="Año de explotación",
            yaxis_title="Diferencia proyectada UF",
            title_x=0.02,
        )
        st.plotly_chart(fig2, use_container_width=True)

        st.markdown("---")
        st.markdown("## 📌 Comentarios automáticos por año")

        comentarios = []
        for _, row in tabla_anios.iterrows():
            anio = row.get("Año", "")
            dif = float(row.get("Diferencia proyectada", 0) or 0)
            rev = float(row.get("En revisión desembolsos", 0) or 0)
            desemb = float(row.get("Desembolsos explotación", 0) or 0)
            vma = float(row.get("VMA año explotación", 0) or 0)

            if dif < 0:
                texto = "Presenta diferencia proyectada negativa; requiere seguimiento financiero y revisión de respaldo."
            elif rev > 0:
                texto = "Presenta montos en revisión; el resultado puede variar según la validación de antecedentes."
            elif desemb > 0 and vma > 0:
                texto = "Presenta desembolsos registrados dentro del período; mantener control documental."
            else:
                texto = "Sin observaciones financieras relevantes para el período."

            comentarios.append({
                "Año": anio,
                "VMA año explotación": uf_fmt(vma),
                "Desembolsos": uf_fmt(desemb),
                "En revisión": uf_fmt(rev),
                "Diferencia proyectada": uf_fmt(dif),
                "Comentario ejecutivo": texto,
            })

        st.dataframe(
            preparar_dataframe_streamlit(pd.DataFrame(comentarios)),
            use_container_width=True,
            hide_index=True,
        )

        with st.expander("📄 Ver planilla original completa"):
            st.dataframe(
                preparar_dataframe_streamlit(df_raw),
                use_container_width=True,
                hide_index=True,
            )

    except Exception as e:
        st.error("No se pudo cargar la Reserva Presupuestaria.")
        st.warning(
            "Verifica que el archivo sea .xlsx, que la hoja exista y que la cuenta de servicio tenga permisos de lectura."
        )
        st.code(str(e))

elif page == "Alertas":
    header("Alertas", "Procesos críticos, atrasados u observados.")
    view = df[(df["Prioridad"].astype(str).str.lower() == "alta") | (df["Vencido"] == True) | (df["Estado"].astype(str).str.lower().str.contains("observ", na=False))] if not df.empty else pd.DataFrame()
    st.dataframe(preparar_dataframe_streamlit(view), use_container_width=True, hide_index=True)

elif page == "Registro":
    header("Registro de Solicitud", "Ingreso directo a Google Sheets.")
    if not can_edit():
        st.error("No tiene permisos para registrar.")
        st.stop()
    with st.form("registro"):
        id_ctar = st.text_input("ID CTAR", value=f"CTAR-{date.today().year}-XXX")
        sic = st.text_input("SIC")
        id_equipo = st.text_input("ID Equipo")
        id_servicio = st.text_input("ID Servicio")
        id_tipo = st.text_input("ID Tipo Proceso")
        id_estado = st.text_input("ID Estado")
        id_resp = st.text_input("ID Responsable")
        id_prioridad = st.text_input("ID Prioridad")
        fecha_ingreso = st.date_input("Fecha ingreso", value=date.today())
        fecha_compromiso = st.date_input("Fecha compromiso", value=date.today())
        motivo = st.text_area("Motivo")
        riesgo = st.text_area("Riesgo clínico")
        ultima = st.text_area("Última gestión")
        proxima = st.text_area("Próxima acción")
        link = st.text_input("Link documento")
        submit = st.form_submit_button("Guardar en Google Sheets")
    if submit:
        row = [id_ctar, sic, id_equipo, id_servicio, id_tipo, id_estado, id_resp, id_prioridad, str(fecha_ingreso), str(fecha_compromiso), "", motivo, riesgo, ultima, proxima, link]
        try:
            append_to_sheet("FACT_CTAR_SEGUIMIENTO", row)
            st.success("Solicitud guardada correctamente.")
            st.rerun()
        except Exception as e:
            st.error("No se pudo guardar en Google Sheets.")
            st.code(str(e))

elif page == "Configuración":
    header("Configuración", "Validación de conexión y estructura Google Sheets.")
    st.markdown("### Hojas detectadas")
    st.write(list(tables.keys()))
    st.markdown("### Hojas faltantes")
    st.write(missing if missing else "No faltan hojas.")
    st.markdown("### Cantidad de registros")
    st.write({k: len(v) for k, v in tables.items()})
